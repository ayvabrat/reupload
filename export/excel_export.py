"""
Экспорт данных из БД в многостраничный Excel (.xlsx).
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, nulls_last, select
from sqlalchemy.orm import Session

import config
from core.models import Channel, SearchSession, Video


def _apply_export_filters(
    q,
    platform: str | None,
    ai: str | None,
    ai_category: str | None,
    duration_filter: str | None,
    profile_id: int | None = None,
):
    """duration_filter: None | missing | present (нет длительности в БД / есть)."""
    if profile_id is not None:
        q = q.where(Video.profile_id == profile_id)
    if platform in ("vk", "rutube"):
        q = q.where(Video.platform == platform)
    if ai == "yes":
        q = q.where(Video.ai_match.is_(True))
    elif ai == "no":
        q = q.where(Video.ai_match.is_(False))
    elif ai == "pending":
        q = q.where(Video.ai_match.is_(None))
    if ai_category in ("reaction", "series", "team"):
        q = q.where(Video.ai_category == ai_category)
    if duration_filter == "missing":
        q = q.where(Video.duration.is_(None))
    elif duration_filter == "present":
        q = q.where(Video.duration.isnot(None))
    return q


def _fmt_dt(dt: datetime | None) -> str:
    if dt is None:
        return ""
    return dt.strftime("%d.%m.%Y %H:%M")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def _header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="366092")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _profile_id_from_filters(xf: dict) -> int | None:
    raw = xf.get("profile_id")
    if isinstance(raw, int):
        return raw
    if isinstance(raw, str) and raw.strip().isdigit():
        return int(raw.strip())
    return None


def export_excel(
    session: Session,
    keywords_used: str,
    out_dir: Path | None = None,
    *,
    export_filters: dict | None = None,
) -> Path:
    """
    Создаёт файл reupload_report_YYYY-MM-DD_HH-MM.xlsx и возвращает путь.
    """
    config.ensure_directories()
    out = out_dir or config.EXPORT_DIR
    out.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    path = out / f"reupload_report_{ts}.xlsx"
    xf0 = export_filters or {}
    pid = _profile_id_from_filters(xf0)

    wb = Workbook()
    # --- Лист 1: Каналы ---
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "Каналы"
    headers1 = [
        "№",
        "Платформа",
        "Название канала",
        "Ссылка на канал",
        "Кол-во совпавших видео",
        "Подписчики",
        "Первое обнаружение",
        "Последнее обнаружение",
        "Подозрительный",
    ]
    for c, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        _header_style(cell)

    ch_q = select(Channel).order_by(Channel.total_matching_videos.desc())
    if pid is not None:
        ch_q = ch_q.where(Channel.profile_id == pid)
    ch_rows = session.execute(ch_q).scalars().all()
    red = PatternFill("solid", fgColor="FFCCCC")
    yellow = PatternFill("solid", fgColor="FFFFCC")
    for i, ch in enumerate(ch_rows, 1):
        row = i + 1
        ws1.cell(row=row, column=1, value=i)
        ws1.cell(row=row, column=2, value=ch.platform.upper())
        ws1.cell(row=row, column=3, value=ch.channel_name)
        link_cell = ws1.cell(row=row, column=4, value=ch.channel_url)
        link_cell.hyperlink = ch.channel_url
        link_cell.style = "Hyperlink"
        ws1.cell(row=row, column=5, value=ch.total_matching_videos)
        ws1.cell(row=row, column=6, value=ch.subscriber_count if ch.subscriber_count is not None else "")
        ws1.cell(row=row, column=7, value=_fmt_dt(ch.first_seen))
        ws1.cell(row=row, column=8, value=_fmt_dt(ch.last_seen))
        ws1.cell(row=row, column=9, value="Да" if ch.is_suspicious else "Нет")
        cnt = ch.total_matching_videos
        fill = None
        if cnt > 10:
            fill = red
        elif cnt >= 5:
            fill = yellow
        if fill:
            for c in range(1, 10):
                ws1.cell(row=row, column=c).fill = fill

    ws1.freeze_panes = "A2"
    _auto_width(ws1)

    # --- Лист 2: Все видео ---
    ws2 = wb.create_sheet("Все видео")
    h2 = [
        "№",
        "Платформа",
        "Канал",
        "Название видео",
        "Ссылка",
        "Просмотры",
        "Длительность",
        "Дата загрузки",
        "Совпавшие ключевые слова",
        "Где совпало",
        "GigaChat (ШГШ)",
        "Тип контента ШГШ",
        "Комментарий ИИ",
    ]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        _header_style(cell)
    pl = xf0.get("platform")
    if pl not in ("vk", "rutube"):
        pl = None
    ai_f = xf0.get("ai")
    if ai_f not in ("yes", "no", "pending"):
        ai_f = None
    cat_f = xf0.get("ai_category")
    if cat_f not in ("reaction", "series", "team"):
        cat_f = None
    dur_f = xf0.get("duration_filter")
    if dur_f not in ("missing", "present"):
        dur_f = None

    q = select(Video, Channel).join(Channel, Video.channel_id == Channel.id)
    q = _apply_export_filters(q, pl, ai_f, cat_f, dur_f, pid)
    q = q.order_by(Video.found_date.desc())
    for idx, (v, ch) in enumerate(session.execute(q).all(), 1):
        r = idx + 1
        ws2.cell(row=r, column=1, value=idx)
        ws2.cell(row=r, column=2, value=v.platform.upper())
        ws2.cell(row=r, column=3, value=ch.channel_name)
        ws2.cell(row=r, column=4, value=v.title)
        lc = ws2.cell(row=r, column=5, value=v.video_url)
        lc.hyperlink = v.video_url
        lc.style = "Hyperlink"
        ws2.cell(row=r, column=6, value=v.views if v.views is not None else "")
        ws2.cell(row=r, column=7, value=v.duration if v.duration is not None else "")
        ws2.cell(row=r, column=8, value=_fmt_dt(v.upload_date))
        ws2.cell(row=r, column=9, value=v.matched_keywords)
        ws2.cell(row=r, column=10, value=v.match_location)
        ai_txt = (
            "Да"
            if v.ai_match is True
            else ("Нет" if v.ai_match is False else "—")
        )
        ws2.cell(row=r, column=11, value=ai_txt)
        cat_txt = {
            "reaction": "Реакции / перезаливы",
            "series": "Серии сериала",
            "team": "Команда ШГШ",
        }.get(v.ai_category or "", "")
        ws2.cell(row=r, column=12, value=cat_txt)
        ws2.cell(row=r, column=13, value=v.ai_note or "")
    ws2.freeze_panes = "A2"
    _auto_width(ws2)

    # --- Лист 3: По каналам (подозрительные) ---
    ws3 = wb.create_sheet("По каналам")
    row = 1
    susp = [c for c in ch_rows if c.total_matching_videos > 5 or c.is_suspicious]
    for ch in sorted(susp, key=lambda x: x.total_matching_videos, reverse=True):
        ws3.cell(row=row, column=1, value=f'--- Канал: "{ch.channel_name}" ({ch.platform.upper()}) ---')
        row += 1
        ws3.cell(row=row, column=1, value=f"--- Ссылка: {ch.channel_url} ---")
        lc = ws3.cell(row=row, column=1)
        lc.hyperlink = ch.channel_url
        lc.style = "Hyperlink"
        row += 1
        vq = select(Video).where(Video.channel_id == ch.id)
        if pid is not None:
            vq = vq.where(Video.profile_id == pid)
        vids = session.execute(vq.order_by(nulls_last(Video.views.desc()))).scalars().all()
        ws3.cell(row=row, column=1, value=f"--- Найденных видео: {len(vids)} ---")
        row += 1
        sub_h = ["№", "Название", "Ссылка", "Просмотры", "Дата", "Ключевые слова"]
        for c, h in enumerate(sub_h, 1):
            cell = ws3.cell(row=row, column=c, value=h)
            _header_style(cell)
        row += 1
        for j, v in enumerate(vids, 1):
            ws3.cell(row=row, column=1, value=j)
            ws3.cell(row=row, column=2, value=v.title)
            u = ws3.cell(row=row, column=3, value=v.video_url)
            u.hyperlink = v.video_url
            u.style = "Hyperlink"
            ws3.cell(row=row, column=4, value=v.views if v.views is not None else "")
            ws3.cell(row=row, column=5, value=_fmt_dt(v.upload_date))
            ws3.cell(row=row, column=6, value=v.matched_keywords)
            row += 1
        row += 1
    _auto_width(ws3)

    # --- Лист 4: Статистика ---
    ws4 = wb.create_sheet("Статистика")
    vcount = select(func.count()).select_from(Video)
    chcount = select(func.count()).select_from(Channel)
    aiyes = select(func.count()).select_from(Video).where(Video.ai_match.is_(True))
    bypl = select(Video.platform, func.count(Video.id)).group_by(Video.platform)
    if pid is not None:
        vcount = select(func.count()).select_from(Video).where(Video.profile_id == pid)
        chcount = select(func.count()).select_from(Channel).where(Channel.profile_id == pid)
        aiyes = select(func.count()).select_from(Video).where(
            Video.profile_id == pid, Video.ai_match.is_(True)
        )
        bypl = bypl.where(Video.profile_id == pid)
    total_v = session.scalar(vcount) or 0
    total_ch = session.scalar(chcount) or 0
    ai_yes = session.scalar(aiyes) or 0
    by_pl = dict(session.execute(bypl).all())
    ls_base = select(SearchSession).where(SearchSession.finished_at.isnot(None))
    if pid is not None:
        ls_base = ls_base.where(SearchSession.profile_id == pid)
    last_sess = session.execute(
        ls_base.order_by(SearchSession.finished_at.desc(), SearchSession.id.desc()).limit(1)
    ).scalars().first()

    lines = [
        ("Всего видео в базе", total_v),
        ("Уникальных каналов", total_ch),
        ("VK", by_pl.get("vk", 0)),
        ("Rutube", by_pl.get("rutube", 0)),
        ("GigaChat: релевантно ШГШ/Гладенко", ai_yes),
        ("Дата последнего сканирования", _fmt_dt(last_sess.finished_at) if last_sess else "—"),
        ("Ключевые слова (этот отчёт)", keywords_used),
    ]
    r0 = 1
    for label, val in lines:
        ws4.cell(row=r0, column=1, value=label).font = Font(bold=True)
        ws4.cell(row=r0, column=2, value=val)
        r0 += 1

    ws4.cell(row=r0 + 1, column=1, value="Топ-10 каналов по совпадениям").font = Font(bold=True)
    top_q = select(Channel).order_by(Channel.total_matching_videos.desc()).limit(10)
    if pid is not None:
        top_q = top_q.where(Channel.profile_id == pid)
    top = session.execute(top_q).scalars().all()
    r0 += 2
    for i, ch in enumerate(top, 1):
        ws4.cell(row=r0, column=1, value=f"{i}. {ch.channel_name} ({ch.platform})")
        ws4.cell(row=r0, column=2, value=ch.total_matching_videos)
        r0 += 1
    _auto_width(ws4)

    wb.save(path)
    return path
